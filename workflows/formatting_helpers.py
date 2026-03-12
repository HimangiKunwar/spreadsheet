"""
Formatting Macros Helper Module for SmartSheet Pro
These operations only work on Excel files (.xlsx)
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    numbers, Protection
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


# Color mapping (name to hex)
COLOR_MAP = {
    'red': 'FF0000',
    'green': '00FF00',
    'blue': '0000FF',
    'yellow': 'FFFF00',
    'orange': 'FFA500',
    'purple': '800080',
    'pink': 'FFC0CB',
    'cyan': '00FFFF',
    'gray': '808080',
    'grey': '808080',
    'white': 'FFFFFF',
    'black': '000000',
    'light_green': '90EE90',
    'light_blue': 'ADD8E6',
    'light_yellow': 'FFFFE0',
    'light_red': 'FFCCCB',
}


def get_color_hex(color_input):
    """Convert color name or hex to proper hex format"""
    if not color_input:
        return 'FFFF00'  # Default yellow
    
    color = color_input.lower().strip()
    
    # Check if it's a named color
    if color in COLOR_MAP:
        return COLOR_MAP[color]
    
    # Check if it's already hex (with or without #)
    hex_color = color.replace('#', '').upper()
    if len(hex_color) == 6 and all(c in '0123456789ABCDEF' for c in hex_color):
        return hex_color
    
    return 'FFFF00'  # Default


def get_column_index(col_name, ws):
    """Get column index from column name"""
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == col_name:
            return idx
    return None


def dataframe_to_excel_with_formatting(df, file_path):
    """Convert DataFrame to Excel workbook for formatting operations"""
    wb = Workbook()
    ws = wb.active
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            # Handle NaN/None
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value=None)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    return wb, ws


class FormattingOperations:
    """Class to handle all formatting operations"""
    
    def __init__(self, workbook, worksheet):
        self.wb = workbook
        self.ws = worksheet
        self.max_row = worksheet.max_row
        self.max_col = worksheet.max_column
    
    # 1. Auto-fit Column Width
    def autofit_column_width(self, column=None):
        """Adjust column width based on content"""
        if column and column != 'all':
            col_idx = get_column_index(column, self.ws)
            if col_idx:
                self._autofit_single_column(col_idx)
        else:
            # Auto-fit all columns
            for col_idx in range(1, self.max_col + 1):
                self._autofit_single_column(col_idx)
    
    def _autofit_single_column(self, col_idx):
        """Auto-fit a single column"""
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row in range(1, self.max_row + 1):
            cell = self.ws.cell(row=row, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # Add padding
        adjusted_width = min(max_length + 2, 50)  # Max 50
        self.ws.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Auto-fit Row Height
    def autofit_row_height(self, row_range=None):
        """Adjust row height based on content"""
        if row_range and row_range != 'all':
            # Parse range like "1-10" or single "5"
            if '-' in str(row_range):
                start, end = map(int, row_range.split('-'))
            else:
                start = end = int(row_range)
            
            for row in range(start, end + 1):
                self.ws.row_dimensions[row].height = 15  # Default height
        else:
            for row in range(1, self.max_row + 1):
                self.ws.row_dimensions[row].height = 15
    
    # 3. Apply Bold/Italic
    def apply_bold_italic(self, column, style='bold', row_range=None):
        """Apply bold and/or italic to column"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        is_bold = style in ['bold', 'both']
        is_italic = style in ['italic', 'both']
        
        start_row = 1 if row_range == 'header' else 2
        end_row = 1 if row_range == 'header' else self.max_row
        
        for row in range(start_row, end_row + 1):
            cell = self.ws.cell(row=row, column=col_idx)
            current_font = cell.font
            cell.font = Font(
                name=current_font.name,
                size=current_font.size,
                bold=is_bold,
                italic=is_italic,
                color=current_font.color
            )
    
    # 4. Change Font
    def change_font(self, column, font_name='Calibri'):
        """Change font family for column"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        for row in range(1, self.max_row + 1):
            cell = self.ws.cell(row=row, column=col_idx)
            current_font = cell.font
            cell.font = Font(
                name=font_name,
                size=current_font.size or 11,
                bold=current_font.bold,
                italic=current_font.italic,
                color=current_font.color
            )
    
    # 5. Change Font Size
    def change_font_size(self, column, font_size=11):
        """Change font size for column"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        for row in range(1, self.max_row + 1):
            cell = self.ws.cell(row=row, column=col_idx)
            current_font = cell.font
            cell.font = Font(
                name=current_font.name or 'Calibri',
                size=int(font_size),
                bold=current_font.bold,
                italic=current_font.italic,
                color=current_font.color
            )
    
    # 6. Apply Cell Color (Background)
    def apply_cell_color(self, column, color='yellow', row_condition=None):
        """Apply background color to cells"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        hex_color = get_color_hex(color)
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        
        if row_condition == 'header':
            self.ws.cell(row=1, column=col_idx).fill = fill
        elif row_condition == 'all' or not row_condition:
            for row in range(1, self.max_row + 1):
                self.ws.cell(row=row, column=col_idx).fill = fill
        else:
            # Data rows only
            for row in range(2, self.max_row + 1):
                self.ws.cell(row=row, column=col_idx).fill = fill
    
    # 7. Apply Text Color
    def apply_text_color(self, column, color='red', condition=None, condition_value=None):
        """Apply text/font color to cells"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        hex_color = get_color_hex(color)
        
        for row in range(2, self.max_row + 1):
            cell = self.ws.cell(row=row, column=col_idx)
            apply_color = True
            
            # Check condition if specified
            if condition and condition_value is not None:
                cell_value = cell.value
                try:
                    if condition == 'negative':
                        apply_color = cell_value is not None and float(cell_value) < 0
                    elif condition == 'less_than':
                        apply_color = cell_value is not None and float(cell_value) < float(condition_value)
                    elif condition == 'greater_than':
                        apply_color = cell_value is not None and float(cell_value) > float(condition_value)
                    elif condition == 'equals':
                        apply_color = str(cell_value) == str(condition_value)
                except (ValueError, TypeError):
                    apply_color = False
            
            if apply_color:
                current_font = cell.font
                cell.font = Font(
                    name=current_font.name or 'Calibri',
                    size=current_font.size or 11,
                    bold=current_font.bold,
                    italic=current_font.italic,
                    color=hex_color
                )
    
    # 8. Add Borders
    def add_borders(self, column=None, border_style='thin'):
        """Add borders to cells"""
        style_map = {
            'thin': Side(style='thin'),
            'thick': Side(style='thick'),
            'medium': Side(style='medium'),
            'dashed': Side(style='dashed'),
            'dotted': Side(style='dotted'),
        }
        
        side = style_map.get(border_style, Side(style='thin'))
        border = Border(left=side, right=side, top=side, bottom=side)
        
        if column and column != 'all':
            col_idx = get_column_index(column, self.ws)
            if col_idx:
                for row in range(1, self.max_row + 1):
                    self.ws.cell(row=row, column=col_idx).border = border
        else:
            # Apply to all cells
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    self.ws.cell(row=row, column=col).border = border
    
    # 9. Merge Cells
    def merge_cells(self, cell_range):
        """Merge cells in specified range"""
        try:
            self.ws.merge_cells(cell_range)
        except Exception as e:
            print(f"Error merging cells: {e}")
    
    # 10. Unmerge Cells
    def unmerge_cells(self, cell_range):
        """Unmerge cells in specified range"""
        try:
            self.ws.unmerge_cells(cell_range)
        except Exception as e:
            print(f"Error unmerging cells: {e}")
    
    # 11. Align Text
    def align_text(self, column, horizontal='center', vertical='center'):
        """Align text in cells"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        alignment = Alignment(horizontal=horizontal, vertical=vertical)
        
        for row in range(1, self.max_row + 1):
            self.ws.cell(row=row, column=col_idx).alignment = alignment
    
    # 12. Apply Number Format
    def apply_number_format(self, column, format_type='comma'):
        """Apply number formatting"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        format_map = {
            'comma': '#,##0',
            'comma_decimal': '#,##0.00',
            'currency_usd': '$#,##0.00',
            'currency_inr': '₹#,##0.00',
            'currency_eur': '€#,##0.00',
            'percentage': '0.00%',
            'percentage_whole': '0%',
            'decimal_2': '0.00',
            'decimal_4': '0.0000',
            'scientific': '0.00E+00',
            'date': 'DD/MM/YYYY',
            'date_us': 'MM/DD/YYYY',
        }
        
        number_format = format_map.get(format_type, '#,##0')
        
        for row in range(2, self.max_row + 1):
            self.ws.cell(row=row, column=col_idx).number_format = number_format
    
    # 13. Conditional Formatting
    def conditional_formatting(self, column, condition='less_than', value=0, color='red'):
        """Apply conditional formatting based on cell values"""
        col_idx = get_column_index(column, self.ws)
        if not col_idx:
            return
        
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{self.max_row}"
        
        hex_color = get_color_hex(color)
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        
        operator_map = {
            'less_than': 'lessThan',
            'greater_than': 'greaterThan',
            'equal': 'equal',
            'not_equal': 'notEqual',
            'less_than_or_equal': 'lessThanOrEqual',
            'greater_than_or_equal': 'greaterThanOrEqual',
            'between': 'between',
        }
        
        op = operator_map.get(condition, 'lessThan')
        
        rule = CellIsRule(
            operator=op,
            formula=[str(value)],
            fill=fill
        )
        
        self.ws.conditional_formatting.add(cell_range, rule)
    
    # 14. Copy Formatting
    def copy_formatting(self, source_column, target_column):
        """Copy formatting from one column to another"""
        source_idx = get_column_index(source_column, self.ws)
        target_idx = get_column_index(target_column, self.ws)
        
        if not source_idx or not target_idx:
            return
        
        for row in range(1, self.max_row + 1):
            source_cell = self.ws.cell(row=row, column=source_idx)
            target_cell = self.ws.cell(row=row, column=target_idx)
            
            # Copy font
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # Copy fill
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # Copy border
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            
            # Copy alignment
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical
                )
            
            # Copy number format
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
    
    # 15. Clear Formatting
    def clear_formatting(self, column=None):
        """Remove all formatting from column or entire sheet"""
        default_font = Font(name='Calibri', size=11)
        default_fill = PatternFill(fill_type=None)
        default_border = Border()
        default_alignment = Alignment()
        
        if column and column != 'all':
            col_idx = get_column_index(column, self.ws)
            if col_idx:
                for row in range(1, self.max_row + 1):
                    cell = self.ws.cell(row=row, column=col_idx)
                    cell.font = default_font
                    cell.fill = default_fill
                    cell.border = default_border
                    cell.alignment = default_alignment
                    cell.number_format = 'General'
        else:
            # Clear all
            for row in range(1, self.max_row + 1):
                for col in range(1, self.max_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.font = default_font
                    cell.fill = default_fill
                    cell.border = default_border
                    cell.alignment = default_alignment
                    cell.number_format = 'General'